import json
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side, PatternFill

def readJson(filePath):
    with open(filePath, 'r') as file:
        return json.load(file)

def writeJson(filePath, data):
    with open(filePath, 'w') as file:
        json.dump(data, file, indent=4)

def writeToExcel(data, outputFile):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Enriched Data"

    headers = [
        "Full Name", "First Name", "Last Name", "Position", "Company Name", "Company URL", "LinkedIn", "Email 1", "Email 2", "Location", "Phone"
    ]
    
    header_font = Font(bold=True, color="000000", size=12)
    header_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    for colNum, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=colNum, value=header)
        cell.font = header_font
        cell.fill = header_fill

    for rowNum, entry in enumerate(data, 2):
        ws.cell(row=rowNum, column=1, value=entry['fullName'])
        ws.cell(row=rowNum, column=2, value=entry['firstName'])
        ws.cell(row=rowNum, column=3, value=entry['lastName'])
        ws.cell(row=rowNum, column=4, value=entry['companyPosition'])
        ws.cell(row=rowNum, column=5, value=entry['companyName'] or entry['company'])
        
        ws.cell(row=rowNum, column=6, value=entry['companyUrl']).hyperlink = entry['companyUrl']
        ws.cell(row=rowNum, column=6).font = Font(color="0000FF", underline="single") if entry['companyUrl'] else Font()
        
        ws.cell(row=rowNum, column=7, value=entry['currentUrl']).hyperlink = entry['currentUrl']
        ws.cell(row=rowNum, column=7).font = Font(color="0000FF", underline="single") if entry['currentUrl'] else Font()
        
        ws.cell(row=rowNum, column=8, value=entry['email0'] or '').hyperlink = f'mailto:{entry["email0"]}' if entry['email0'] else ''
        ws.cell(row=rowNum, column=8).font = Font(color="0000FF", underline="single") if entry['email0'] else Font()
        
        ws.cell(row=rowNum, column=9, value=entry['email1'] or '').hyperlink = f'mailto:{entry["email1"]}' if entry['email1'] else ''
        ws.cell(row=rowNum, column=9).font = Font(color="0000FF", underline="single") if entry['email1'] else Font()
        
        ws.cell(row=rowNum, column=10, value=entry['companyLocation'])
        ws.cell(row=rowNum, column=11, value=entry['phone'] or '')

    ws.column_dimensions[get_column_letter(1)].width = max(len(entry['fullName']) for entry in data) + 2
    ws.column_dimensions[get_column_letter(2)].width = 15
    ws.column_dimensions[get_column_letter(3)].width = 15
    ws.column_dimensions[get_column_letter(4)].width = 25
    ws.column_dimensions[get_column_letter(5)].width = 35
    ws.column_dimensions[get_column_letter(6)].width = 15
    ws.column_dimensions[get_column_letter(7)].width = 15
    ws.column_dimensions[get_column_letter(8)].width = 20
    ws.column_dimensions[get_column_letter(9)].width = 15
    ws.column_dimensions[get_column_letter(10)].width = 45
    ws.column_dimensions[get_column_letter(11)].width = 15

    thin = Side(border_style="thin", color="000000")
    for row in ws.iter_rows(min_row=1, max_col=len(headers), max_row=len(data)+1):
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    dataToSort = list(ws.iter_rows(min_row=2, max_row=len(data)+1, values_only=True))
    dataToSort.sort(key=lambda x: (x[6] == '', x[6]))

    for rowNum, rowData in enumerate(dataToSort, start=2):
        for colNum, value in enumerate(rowData, start=1):
            ws.cell(row=rowNum, column=colNum, value=value)

    wb.save(outputFile)

def processJson(filePath):
    data = readJson(filePath)

    formattedData = []
    for entry in data:
        formattedEntry = {
            'fullName': entry['fullName'],
            'companyPosition': entry.get('companyPosition', ''),
            'currentUrl': entry.get('currentUrl', ''),
            'email0': entry.get('email0', ''),
            'email1': entry.get('email1', ''),
            'phone': entry.get('phone', ''),
            'companyName': entry.get('companyName', ''),
            'companyUrl': entry.get('companyUrl', ''),
            'companyLocation': entry.get('companyLocation', ''),
            'firstName': entry.get('firstName', ''),
            'lastName': entry.get('lastName', ''),
            'company': entry.get('company', '')
        }

        formattedData.append(formattedEntry)

    outputFile = 'contacts.xlsx'
    writeToExcel(formattedData, outputFile)
    print(f"Data has been written to {outputFile}")

filePath = 'output.json'
processJson(filePath)
