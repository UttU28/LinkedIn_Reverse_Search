import re
import pandas as pd
from openpyxl import load_workbook

def splitFullName(fullName):
    prefixPattern = r'^(Dr\.?|H\.E\.?\.?)\s+'
    match = re.match(prefixPattern, fullName)
    
    if match:
        namePart = fullName[match.end():].strip()
    else:
        namePart = fullName.strip()
    
    nameParts = re.split(r'(?<!^)\s+(?=[A-Z])|(?<=[a-z])(?=[A-Z])', namePart)
    
    if len(nameParts) > 1:
        firstName = ' '.join(nameParts[:-1])
        lastName = nameParts[-1]
    else:
        firstName = namePart
        lastName = ''
    
    return {
        'firstName': firstName.strip(),
        'lastName': lastName.strip()
    }

filePath = 'People.xlsx'
df = pd.read_excel(filePath)

fullNameList = df.columns[0]
fullNames = df[fullNameList].to_list()

firstNames = []
lastNames = []

if 'First Name' not in df.columns and 'Last Name' not in df.columns:
    for fullName in fullNames:
        result = splitFullName(fullName)
        firstNames.append(result['firstName'])
        lastNames.append(result['lastName'])

    df['First Name'] = firstNames
    df['Last Name'] = lastNames

for col in ['LinkedIn', 'Email', 'Company URL']:
    if col not in df.columns:
        df[col] = ''

newOrder = ['First Name', 'Last Name', 'Company', 'Title', 'Full Name', 'LinkedIn', 'Email', 'Company URL']
existingOrder = [col for col in newOrder if col in df.columns]
df = df[existingOrder]

outputFilePath = 'People2.xlsx'
df.to_excel(outputFilePath, index=False)

wb = load_workbook(outputFilePath)
ws = wb.active

for col in ['A', 'B', 'C', 'D']:
    maxLength = 0
    column = ws[col]
    for cell in column:
        try:
            if len(str(cell.value)) > maxLength:
                maxLength = len(str(cell.value))
        except:
            pass
    adjustedWidth = maxLength + 2
    ws.column_dimensions[col].width = adjustedWidth

ws.column_dimensions['C'].width = 45 
ws.column_dimensions['D'].width = 45 
ws.column_dimensions['E'].width = 45 
ws.column_dimensions['F'].width = 20 
ws.column_dimensions['G'].width = 20 
ws.column_dimensions['H'].width = 20 

wb.save(outputFilePath)

print("Excel file updated with specified column order and adjusted widths.")
